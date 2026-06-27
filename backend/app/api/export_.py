from __future__ import annotations

import csv
import io
import json
from typing import Annotated, Literal, Optional
from uuid import UUID

import openpyxl
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

from app.core.db import get_db

router = APIRouter(tags=["export"])

_RU = {
    'shortcode': 'Код',
    'url': 'Ссылка',
    'type': 'Тип',
    'caption': 'Текст поста',
    'author_handle': 'Автор',
    'author_followers': 'Подписчики',
    'views': 'Просмотры',
    'likes': 'Лайки',
    'comments': 'Комменты',
    'posted_at': 'Дата',
    'er': 'Залётность',
    'lpf': 'Лайки/подп %',
    'cpf': 'Комм/подп %',
    'eng': 'Вовлечённость %',
    'transcript': 'Расшифровка',
    'transcript_ru': 'Расшифровка RU',
    'note': 'Заметка',
}


def _fetch(ids: Optional[list[UUID]]) -> list[dict]:
    db = get_db()
    q = db.table('reels').select('*, transcripts(text, text_ru, status), reel_notes(note)')
    if ids:
        q = q.in_('id', [str(i) for i in ids])
    return q.execute().data


def _to_flat(r: dict, lang: str) -> dict:
    t = (r.get('transcripts') or [{}])[0]
    note_list = r.get('reel_notes') or []
    note = note_list[0].get('note', '') if note_list else ''
    row = {
        'shortcode': r.get('shortcode', ''),
        'url': r.get('url', ''),
        'type': r.get('type', ''),
        'caption': r.get('caption', ''),
        'author_handle': r.get('author_handle', ''),
        'author_followers': r.get('author_followers', ''),
        'views': r.get('views', ''),
        'likes': r.get('likes', ''),
        'comments': r.get('comments', ''),
        'posted_at': str(r.get('posted_at') or ''),
        'er': r.get('er', ''),
        'lpf': r.get('lpf', ''),
        'cpf': r.get('cpf', ''),
        'eng': r.get('eng', ''),
        'note': note,
    }
    if lang == 'orig':
        row['transcript'] = t.get('text', '')
    elif lang == 'ru':
        row['transcript'] = t.get('text_ru') or t.get('text', '')
    else:
        row['transcript'] = t.get('text', '')
        row['transcript_ru'] = t.get('text_ru', '')
    return row


@router.get("/export")
async def export_reels(
    format: Annotated[Literal["csv", "xlsx", "json", "md"], Query()] = "csv",  # noqa: A002
    lang: Annotated[Literal["orig", "ru", "both"], Query()] = "ru",
    scope: Annotated[Literal["all", "selected"], Query()] = "all",
    ids: Optional[list[UUID]] = Query(None),
) -> Response:
    rows = _fetch(ids if scope == 'selected' else None)

    if format == 'csv':
        return _as_csv(rows, lang)
    if format == 'xlsx':
        return _as_xlsx(rows, lang)
    if format == 'json':
        return _as_json(rows, lang)
    raise HTTPException(501, f"формат {format} ещё не реализован")


def _as_csv(rows: list[dict], lang: str) -> StreamingResponse:
    fields = list(_to_flat({}, lang).keys())
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writerow({f: _RU.get(f, f) for f in fields})
    for r in rows:
        writer.writerow(_to_flat(r, lang))
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename="reelscribe.csv"'},
    )


def _as_xlsx(rows: list[dict], lang: str) -> Response:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ReelScribe'
    fields = list(_to_flat({}, lang).keys())
    header_fill = openpyxl.styles.PatternFill('solid', fgColor='1a1a2e')
    header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    ws.append([_RU.get(f, f) for f in fields])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        flat = _to_flat(r, lang)
        ws.append([flat.get(f, '') for f in fields])
    # авто-ширина колонок
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reelscribe.xlsx"'},
    )


def _as_json(rows: list[dict], lang: str) -> Response:
    data = [_to_flat(r, lang) for r in rows]
    return Response(
        content=json.dumps(data, ensure_ascii=False, indent=2),
        media_type='application/json',
        headers={'Content-Disposition': 'attachment; filename="reelscribe.json"'},
    )
